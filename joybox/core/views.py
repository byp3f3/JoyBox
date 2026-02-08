from rest_framework import generics, filters, status, viewsets
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.authtoken.models import Token
from rest_framework.response import Response
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.exceptions import PermissionDenied, ValidationError as DRFValidationError
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth import authenticate
from django.db import models, transaction, IntegrityError, connection
from django.db.models import Q, Sum, Count, Avg, F
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
from django.http import HttpResponse
from datetime import datetime, timedelta
import csv
import io
import json as _json
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
    print(">>> OPENPYXL LOADED SUCCESSFULLY <<<")
except Exception as e:
    OPENPYXL_AVAILABLE = False
    import traceback
    print(f">>> OPENPYXL IMPORT ERROR: {type(e).__name__}: {e} <<<")
    traceback.print_exc()
from django.shortcuts import render, redirect
from django.utils import timezone
from django.conf import settings
from django.http import FileResponse
from decimal import Decimal
from pathlib import Path
import logging
import subprocess
from .models import Product, Category, Brand, ProductImage, ProductAttribute, Review, Wishlist, ParentChild, User, Order, OrderItem, OrderStatus, Address, Role, AuditLog, Cart

logger = logging.getLogger(__name__)
from .audit import log_audit, model_to_log_dict, get_pk, set_audit_user
from .serializers import (
    ProductListSerializer, 
    ProductDetailSerializer,
    CategorySerializer,
    BrandSerializer,
    ReviewSerializer,
    ReviewCreateSerializer,
    ReviewUpdateSerializer,
    UserSerializer,
    UserProfileSerializer,
    UserRegistrationSerializer,
    LoginSerializer,
    WishlistSerializer,
    WishlistCreateSerializer,
    ChildAccountSerializer,
    ChildAccountCreateSerializer,
    ChildAccountUpdateSerializer,
    CartItemSerializer,
    CartAddSerializer,
    CartUpdateSerializer,
    AddressBriefSerializer,
    AddressCreateSerializer,
    CheckoutCreateSerializer,
    OrderSerializer,
    OrderDetailSerializer,
    UserOrderSerializer,
    UserOrderDetailSerializer,
    PaymentProcessSerializer,
    OrderStatusSerializer,
    AuditLogSerializer,
    AdminReviewSerializer,
    ProductCreateUpdateSerializer,
    ProductImageSerializer,
    ProductAttributeSerializer,
    ProductWithRelationsSerializer,
    UserCreateUpdateSerializer,
    RoleSerializer
)
from .filters import ProductFilter
from django.shortcuts import render
from django.db.models import Count, Avg
from django.conf import settings
from django.core.files.storage import default_storage
import os
import uuid

class CategoryListView(generics.ListAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer

class BrandListView(generics.ListAPIView):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer

class ProductListView(generics.ListAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductListSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ProductFilter
    search_fields = ['productName', 'productDescription']
    ordering_fields = ['price', 'productName', 'createdAt']
    ordering = ['productName']

class PopularProductsListView(generics.ListAPIView):
    serializer_class = ProductListSerializer
    
    def get_queryset(self):
        return Product.objects.filter(
            review__isnull=False
        ).annotate(
            avg_rating=Avg('review__rating'),
            review_count=Count('review')
        ).order_by('-avg_rating', '-review_count')[:12]

class ProductDetailView(generics.RetrieveAPIView):
    serializer_class = ProductDetailSerializer

    def get_queryset(self):
        return Product.objects.all().select_related('categoryId', 'brandId').prefetch_related('productimage_set', 'productattribute_set')

class ProductReviewsListView(generics.ListAPIView):
    serializer_class = ReviewSerializer
    
    def get_queryset(self):
        product_id = self.kwargs['product_id']
        return Review.objects.filter(productId=product_id).select_related('userId')

class UserRegistrationView(generics.CreateAPIView):
    serializer_class = UserRegistrationSerializer
    permission_classes = [AllowAny]  
    
    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            user = serializer.save()
            token, created = Token.objects.get_or_create(user=user)
            return Response({
                'token': token.key,
                'user': UserSerializer(user).data
            }, status=status.HTTP_201_CREATED)
        except DRFValidationError:
            raise  # Обрабатывается централизованным обработчиком
        except IntegrityError:
            raise  # Обрабатывается централизованным обработчиком
        except Exception as e:
            logger.exception("Ошибка регистрации")
            return Response({'detail': 'Ошибка при регистрации. Проверьте введённые данные.'}, status=status.HTTP_400_BAD_REQUEST)

class LoginView(ObtainAuthToken):
    serializer_class = LoginSerializer
    permission_classes = [AllowAny]  
    
    def post(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(data=request.data,
                                               context={'request': request})
            serializer.is_valid(raise_exception=True)
            user = serializer.validated_data['user']
            token, created = Token.objects.get_or_create(user=user)
            return Response({
                'token': token.key,
                'user': UserSerializer(user).data
            })
        except DRFValidationError:
            raise  # Обрабатывается централизованным обработчиком
        except Exception as e:
            logger.exception("Ошибка авторизации")
            return Response({'detail': 'Ошибка при входе. Проверьте email и пароль.'}, status=status.HTTP_400_BAD_REQUEST)

    def dispatch(self, request, *args, **kwargs):
        try:
            return super().dispatch(request, *args, **kwargs)
        except DRFValidationError:
            raise
        except Exception as e:
            logger.exception("Ошибка авторизации (dispatch)")
            return Response({'detail': 'Ошибка при входе. Попробуйте позже.'}, status=400)

class UserProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated]  
    
    def get_object(self):
        return self.request.user

def info_view(request):
    if request.user.is_authenticated:
        if hasattr(request.user, 'roleId') and request.user.roleId.roleName in ['Администратор', 'Менеджер']:
            return redirect('/admin-panel/')
    return render(request, 'info.html')

def catalog_page(request):
    if request.user.is_authenticated:
        if hasattr(request.user, 'roleId') and request.user.roleId.roleName in ['Администратор', 'Менеджер']:
            return redirect('/admin-panel/')
    return render(request, 'catalog.html')

def product_detail_page(request, pk):
    if request.user.is_authenticated:
        if hasattr(request.user, 'roleId') and request.user.roleId.roleName in ['Администратор', 'Менеджер']:
            return redirect('/admin-panel/')
    return render(request, 'product_detail.html', {'product_id': pk})

def profile_page(request):
    return render(request, 'profile.html')

def wishlist_page(request):
    return render(request, 'wishlist.html')


def cart_page(request):
    return render(request, 'cart.html')


def checkout_page(request):
    return render(request, 'checkout.html')


def payment_page(request):
    return render(request, 'payment.html')


def orders_page(request):
    return render(request, 'orders.html')


def privacy_page(request):
    return render(request, 'privacy.html')

def admin_panel_page(request):
    return render(request, 'admin_panel.html')

def login_page(request):
    return render(request, 'login.html')

def register_page(request):
    return render(request, 'register.html')

class WishlistListView(generics.ListAPIView):
    serializer_class = WishlistSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName == 'Покупатель':
            children = User.objects.filter(child_relations__userId=user)
            return Wishlist.objects.filter(
                models.Q(userId=user) | models.Q(userId__in=children)
            ).select_related('userId', 'productId')
        else:
            return Wishlist.objects.filter(userId=user).select_related('userId', 'productId')


class WishlistCreateView(generics.CreateAPIView):
    serializer_class = WishlistCreateSerializer
    permission_classes = [IsAuthenticated]
    
    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        if 'productId' not in data and 'product_id' in data:
            data['productId'] = data['product_id']
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(
            WishlistSerializer(serializer.instance).data,
            status=status.HTTP_201_CREATED,
        )


class WishlistDestroyView(generics.DestroyAPIView):
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return Wishlist.objects.filter(userId=self.request.user)


def _cart_allowed(user):
    return hasattr(user, 'roleId') and user.roleId.roleName == 'Покупатель'


class CartListView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Cart.objects.filter(userId=self.request.user).select_related('productId')

    def get(self, request, *args, **kwargs):
        if not _cart_allowed(request.user):
            return Response({'detail': 'Корзина доступна только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        items = self.get_queryset()
        from decimal import Decimal
        total = Decimal('0.00')
        for cart_item in items:
            total += cart_item.productId.price * cart_item.quantity
        serializer = CartItemSerializer(items, many=True)
        return Response({'items': serializer.data, 'total': str(total)})

    def post(self, request, *args, **kwargs):
        if not _cart_allowed(request.user):
            return Response({'detail': 'Корзина доступна только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        data = request.data.copy()
        if 'product_id' in data and 'productId' not in data:
            data['productId'] = data['product_id']
        serializer = CartAddSerializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        try:
            cart_item = serializer.save()
        except DRFValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
        out = CartItemSerializer(cart_item)
        return Response(out.data, status=status.HTTP_201_CREATED)


class CartItemDetailView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Cart.objects.filter(userId=self.request.user).select_related('productId')

    def get_object(self):
        return generics.get_object_or_404(self.get_queryset(), cartId=self.kwargs['pk'])

    def patch(self, request, pk, *args, **kwargs):
        if not _cart_allowed(request.user):
            return Response({'detail': 'Корзина доступна только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        cart_item = self.get_object()
        serializer = CartUpdateSerializer(cart_item, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        out = CartItemSerializer(cart_item)
        return Response(out.data)

    def delete(self, request, pk, *args, **kwargs):
        if not _cart_allowed(request.user):
            return Response({'detail': 'Корзина доступна только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        cart_item = self.get_object()
        cart_item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


def _buyer_only(user):
    return hasattr(user, 'roleId') and user.roleId.roleName == 'Покупатель'


class UserAddressListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = AddressBriefSerializer

    def get_queryset(self):
        if not _buyer_only(self.request.user):
            return Address.objects.none()
        return Address.objects.filter(userId=self.request.user)


class UserAddressCreateView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = AddressCreateSerializer

    def create(self, request, *args, **kwargs):
        if not _buyer_only(request.user):
            return Response({'detail': 'Доступно только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        return super().create(request, *args, **kwargs)


class UserAddressDeleteView(generics.DestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = AddressBriefSerializer

    def get_queryset(self):
        if not _buyer_only(self.request.user):
            return Address.objects.none()
        return Address.objects.filter(userId=self.request.user)

    def destroy(self, request, *args, **kwargs):
        if not _buyer_only(request.user):
            return Response({'detail': 'Доступно только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# Пункты выдачи СДЭК (статика; при необходимости заменить на вызов API СДЭК)
SDEK_PICKUP_POINTS = [
    {'code': 'MSK001', 'name': 'СДЭК ул. Тверская', 'city': 'Москва', 'street': 'ул. Тверская', 'house': '1', 'index': '125009', 'lat': 55.7558, 'lng': 37.6173},
    {'code': 'MSK002', 'name': 'СДЭК Арбат', 'city': 'Москва', 'street': 'ул. Арбат', 'house': '12', 'index': '119002', 'lat': 55.7522, 'lng': 37.5916},
    {'code': 'MSK003', 'name': 'СДЭК Лубянка', 'city': 'Москва', 'street': 'ул. Лубянка', 'house': '5', 'index': '101000', 'lat': 55.7602, 'lng': 37.6256},
    {'code': 'MSK004', 'name': 'СДЭК Комсомольская', 'city': 'Москва', 'street': 'Комсомольская пл.', 'house': '2', 'index': '107140', 'lat': 55.7732, 'lng': 37.6554},
    {'code': 'MSK005', 'name': 'СДЭК Театральный', 'city': 'Москва', 'street': 'Театральный проезд', 'house': '5', 'index': '109012', 'lat': 55.7580, 'lng': 37.6185},
    {'code': 'MSK006', 'name': 'СДЭК Кузнецкий мост', 'city': 'Москва', 'street': 'ул. Кузнецкий Мост', 'house': '21/5', 'index': '107031', 'lat': 55.7619, 'lng': 37.6242},
    {'code': 'MSK007', 'name': 'СДЭК Маяковская', 'city': 'Москва', 'street': 'Триумфальная пл.', 'house': '2', 'index': '125047', 'lat': 55.7704, 'lng': 37.5955},
    {'code': 'MSK008', 'name': 'СДЭК Павелецкая', 'city': 'Москва', 'street': 'Павелецкая пл.', 'house': '1', 'index': '115054', 'lat': 55.7314, 'lng': 37.6361},
]


class CheckoutDeliveryOptionsView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        options = [{'value': value, 'label': label} for value, label in Order.DELIVERY_TYPE_CHOICES]
        return Response(options)


class CheckoutSdekPointsView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        return Response(SDEK_PICKUP_POINTS)


class CheckoutPaymentOptionsView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        delivery_type = request.query_params.get('deliveryType', '')
        if delivery_type == Order.DELIVERY_POINT:
            options = [{'value': Order.PAYMENT_ONLINE, 'label': 'Онлайн'}]
        else:
            options = [{'value': value, 'label': label} for value, label in Order.PAYMENT_TYPE_CHOICES]
        return Response(options)


class CreateOrderView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = CheckoutCreateSerializer

    def post(self, request, *args, **kwargs):
        if not _buyer_only(request.user):
            return Response({'detail': 'Оформление заказов доступно только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        serializer = CheckoutCreateSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        user = request.user

        cart_items = list(Cart.objects.filter(userId=user).select_related('productId'))
        if not cart_items:
            return Response({'detail': 'Корзина пуста. Добавьте товары перед оформлением заказа.'}, status=status.HTTP_400_BAD_REQUEST)

        # Проверка остатков до входа в транзакцию
        for item in cart_items:
            if item.quantity > item.productId.quantity:
                return Response({
                    'detail': f'Недостаточно товара «{item.productId.productName}» на складе (доступно: {item.productId.quantity}).'
                }, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # Устанавливаем ID пользователя для триггеров аудита
                set_audit_user(user)

                # Адрес: самовывоз — фиксированный; пункт выдачи/курьер — из запроса
                if data['deliveryType'] == Order.DELIVERY_PICKUP:
                    address = Address.objects.create(
                        userId=user,
                        city='Москва',
                        street='Театральный проезд',
                        house='5с1',
                        flat='',
                        index='109012'
                    )
                elif data.get('addressId'):
                    address = data['addressId']
                    if address.userId_id != user.userId:
                        raise DRFValidationError({'addressId': 'Указан чужой адрес.'})
                else:
                    na = data['newAddress']
                    address = Address.objects.create(
                        userId=user,
                        city=na['city'],
                        street=na['street'],
                        house=na['house'],
                        flat=na.get('flat') or '',
                        index=na['index']
                    )

                # Вызов хранимой процедуры sp_create_order_from_cart:
                # создаёт заказ, переносит позиции из корзины, уменьшает остатки, очищает корзину
                with connection.cursor() as cursor:
                    cursor.execute(
                        'CALL sp_create_order_from_cart(%s, %s, %s, %s)',
                        [user.userId, address.addressId, data['deliveryType'], data['paymentType']]
                    )

                # Получаем созданный заказ
                order = Order.objects.filter(userId=user).order_by('-createdAt').first()

        except DRFValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
        except IntegrityError as e:
            logger.exception("CreateOrder IntegrityError")
            return Response(
                {'detail': 'Ошибка целостности данных при оформлении заказа. Проверьте адрес и повторите попытку.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.exception("CreateOrder failed")
            # Извлекаем сообщение из ошибки хранимой процедуры
            error_msg = str(e)
            detail = 'Не удалось оформить заказ.'
            if 'Корзина пользователя пуста' in error_msg:
                detail = 'Корзина пуста. Добавьте товары перед оформлением заказа.'
            elif 'Недостаточно товара' in error_msg:
                detail = error_msg.split('\n')[0] if '\n' in error_msg else error_msg
            elif getattr(settings, 'DEBUG', False):
                detail += ' Ошибка: ' + error_msg
            return Response(
                {'detail': detail},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            'orderId': order.orderId,
            'total': str(order.total),
            'message': 'Заказ успешно оформлен.'
        }, status=status.HTTP_201_CREATED)


class UserOrdersListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserOrderSerializer

    def get_queryset(self):
        if not _buyer_only(self.request.user):
            return Order.objects.none()
        return Order.objects.filter(userId=self.request.user).select_related('orderStatusId').order_by('-createdAt')


class UserOrderDetailView(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserOrderDetailSerializer

    def get_queryset(self):
        if not _buyer_only(self.request.user):
            return Order.objects.none()
        return Order.objects.filter(userId=self.request.user).select_related('orderStatusId', 'addressId').prefetch_related('orderitem_set__productId')


class UserOrderCancelView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserOrderDetailSerializer

    def get_queryset(self):
        if not _buyer_only(self.request.user):
            return Order.objects.none()
        return Order.objects.filter(userId=self.request.user).select_related('orderStatusId')

    def post(self, request, pk, *args, **kwargs):
        try:
            # Проверяем, что заказ принадлежит пользователю
            order = self.get_queryset().get(orderId=pk)

            with transaction.atomic():
                # Устанавливаем ID пользователя для триггеров аудита
                set_audit_user(request.user)

                # Вызов хранимой процедуры sp_cancel_order:
                # проверяет статус, возвращает товары на склад, меняет статус и оплату
                with connection.cursor() as cursor:
                    cursor.execute('CALL sp_cancel_order(%s)', [pk])

            order.refresh_from_db()
        except Order.DoesNotExist:
            return Response({'detail': 'Заказ не найден.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            error_msg = str(e)
            if 'уже отменён' in error_msg:
                return Response({'detail': 'Заказ уже отменён.'}, status=status.HTTP_400_BAD_REQUEST)
            elif 'Нельзя отменить' in error_msg or 'доставленный' in error_msg:
                return Response({'detail': error_msg.split('\n')[0]}, status=status.HTTP_400_BAD_REQUEST)
            elif 'не найден' in error_msg:
                return Response({'detail': 'Заказ не найден.'}, status=status.HTTP_404_NOT_FOUND)
            logger.exception("OrderCancel failed")
            return Response({'detail': 'Не удалось отменить заказ.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        serializer = self.get_serializer(order)
        return Response(serializer.data, status=status.HTTP_200_OK)


def _user_has_purchased_product(user, product_id):
    return OrderItem.objects.filter(
        orderId__userId=user,
        productId_id=product_id
    ).exclude(
        Q(orderId__orderStatusId__orderStatusName__iexact='отменен') |
        Q(orderId__orderStatusId__orderStatusName__iexact='отменён')
    ).exists()


class UserReviewListCreateView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ReviewSerializer

    def get(self, request, *args, **kwargs):
        if not _buyer_only(request.user):
            return Response({'detail': 'Доступно только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        product_id = request.query_params.get('product_id')
        if not product_id:
            return Response({'detail': 'Укажите product_id.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            product_id = int(product_id)
        except (TypeError, ValueError):
            return Response({'detail': 'Некорректный product_id.'}, status=status.HTTP_400_BAD_REQUEST)
        reviews = Review.objects.filter(productId_id=product_id, userId=request.user).order_by('-createdAt')[:1]
        serializer = ReviewSerializer(reviews, many=True)
        return Response(serializer.data)

    def post(self, request, *args, **kwargs):
        if not _buyer_only(request.user):
            return Response({'detail': 'Отзывы могут оставлять только покупатели.'}, status=status.HTTP_403_FORBIDDEN)
        serializer = ReviewCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        product = serializer.validated_data['productId']
        product_id = product.productId
        if not _user_has_purchased_product(request.user, product_id):
            return Response(
                {'detail': 'Отзыв можно оставить только на товар из вашего заказа.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if Review.objects.filter(productId_id=product_id, userId=request.user).exists():
            return Response(
                {'detail': 'Вы уже оставили отзыв на этот товар.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        now = timezone.now()
        review = Review.objects.create(
            productId=product,
            userId=request.user,
            rating=serializer.validated_data['rating'],
            reviewText=serializer.validated_data.get('reviewText') or '',
            createdAt=now,
            updatedAt=now
        )
        out = ReviewSerializer(review)
        return Response(out.data, status=status.HTTP_201_CREATED)


def _validate_review_text_profanity(text):
    from .profanity import contains_profanity
    if text and contains_profanity(text):
        from rest_framework.exceptions import ValidationError as DRFValidationError
        raise DRFValidationError({'reviewText': ['Текст отзыва содержит недопустимые выражения. Измените формулировку.']})


class UserReviewDetailView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ReviewSerializer

    def get_queryset(self):
        if not _buyer_only(self.request.user):
            return Review.objects.none()
        return Review.objects.filter(userId=self.request.user).select_related('productId', 'userId')

    def get(self, request, pk, *args, **kwargs):
        review = self.get_queryset().filter(reviewId=pk).first()
        if not review:
            return Response({'detail': 'Отзыв не найден.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(ReviewSerializer(review).data)

    def put(self, request, pk, *args, **kwargs):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk, *args, **kwargs):
        return self._update(request, pk, partial=True)

    def _update(self, request, pk, partial=False):
        if not _buyer_only(request.user):
            return Response({'detail': 'Доступно только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        review = Review.objects.filter(reviewId=pk, userId=request.user).first()
        if not review:
            return Response({'detail': 'Отзыв не найден.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ReviewUpdateSerializer(data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        review.rating = serializer.validated_data.get('rating', review.rating)
        if 'reviewText' in serializer.validated_data:
            review.reviewText = serializer.validated_data['reviewText'] or ''
        review.updatedAt = timezone.now()
        review.save(update_fields=['rating', 'reviewText', 'updatedAt'])
        return Response(ReviewSerializer(review).data)


class PaymentProcessView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PaymentProcessSerializer

    def post(self, request, *args, **kwargs):
        if not _buyer_only(request.user):
            return Response({'detail': 'Оплата доступна только покупателям.'}, status=status.HTTP_403_FORBIDDEN)
        serializer = PaymentProcessSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        order_id = serializer.validated_data['orderId']
        try:
            with transaction.atomic():
                order = Order.objects.select_for_update().get(orderId=order_id, userId=request.user)
                if order.paymentStatus != Order.PAYMENT_STATUS_PENDING:
                    return Response({'detail': 'Заказ уже оплачен или не требует оплаты.'}, status=status.HTTP_400_BAD_REQUEST)
                if order.paymentType != Order.PAYMENT_ONLINE:
                    return Response({'detail': 'Этот заказ не предназначен для онлайн-оплаты.'}, status=status.HTTP_400_BAD_REQUEST)
                order.paymentStatus = Order.PAYMENT_STATUS_PAID
                order.save(update_fields=['paymentStatus'])
        except Order.DoesNotExist:
            return Response({'detail': 'Заказ не найден.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.exception("PaymentProcess failed")
            return Response({'detail': 'Ошибка обработки оплаты. Попробуйте позже.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response({'message': 'Оплата успешно обработана.', 'orderId': order.orderId}, status=status.HTTP_200_OK)


class ParentChildrenListView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return ParentChild.objects.filter(userId=self.request.user).select_related('childId')

    def get(self, request, *args, **kwargs):
        if request.user.roleId.roleName != 'Покупатель':
            return Response({'detail': 'Доступно только покупателям (родителям).'}, status=status.HTTP_403_FORBIDDEN)
        links = self.get_queryset()
        serializer = ChildAccountSerializer(links, many=True)
        return Response(serializer.data)

    def post(self, request, *args, **kwargs):
        if request.user.roleId.roleName != 'Покупатель':
            return Response({'detail': 'Доступно только покупателям (родителям).'}, status=status.HTTP_403_FORBIDDEN)
        serializer = ChildAccountCreateSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        link = serializer.save()
        out = ChildAccountSerializer(link)
        return Response(out.data, status=status.HTTP_201_CREATED)


class ParentChildDetailView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get_link_and_child(self, request, pk):
        link = ParentChild.objects.filter(
            userId=request.user,
            childId_id=pk
        ).select_related('childId').first()
        if not link:
            return None, None
        return link, link.childId

    def get(self, request, pk, *args, **kwargs):
        if request.user.roleId.roleName != 'Покупатель':
            return Response({'detail': 'Доступно только покупателям (родителям).'}, status=status.HTTP_403_FORBIDDEN)
        link, child = self.get_link_and_child(request, pk)
        if not link:
            return Response({'detail': 'Привязанный ребёнок не найден.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ChildAccountSerializer(link)
        return Response(serializer.data)

    def patch(self, request, pk, *args, **kwargs):
        if request.user.roleId.roleName != 'Покупатель':
            return Response({'detail': 'Доступно только покупателям (родителям).'}, status=status.HTTP_403_FORBIDDEN)
        link, child = self.get_link_and_child(request, pk)
        if not link:
            return Response({'detail': 'Привязанный ребёнок не найден.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ChildAccountUpdateSerializer(child, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        out = ChildAccountSerializer(link)
        return Response(out.data)

    def delete(self, request, pk, *args, **kwargs):
        if request.user.roleId.roleName != 'Покупатель':
            return Response({'detail': 'Доступно только покупателям (родителям).'}, status=status.HTTP_403_FORBIDDEN)
        link, _ = self.get_link_and_child(request, pk)
        if not link:
            return Response({'detail': 'Привязанный ребёнок не найден.'}, status=status.HTTP_404_NOT_FOUND)
        link.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AdminPanelView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        user = request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            is_admin = user.roleId.roleName == 'Администратор'
            return Response({
                'message': 'Welcome to admin panel',
                'role': user.roleId.roleName,
                'is_admin': is_admin
            })
        else:
            return Response({
                'message': 'Access denied. Redirecting to main site.',
                'redirect': '/'
            }, status=403)

class AdminDashboardView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        user = request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            total_products = Product.objects.count()
            total_users = User.objects.count()
            total_orders = Order.objects.count()
            total_revenue = sum(order.total for order in Order.objects.all())
            
            return Response({
                'total_products': total_products,
                'total_users': total_users,
                'total_orders': total_orders,
                'total_revenue': float(total_revenue)
            })
        else:
            return Response({'detail': 'Доступ запрещён.'}, status=403)

class AdminProductsView(generics.ListAPIView):
    serializer_class = ProductListSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Product.objects.all().select_related('categoryId', 'brandId')
        else:
            return Product.objects.none()

class AdminProductCreateView(generics.CreateAPIView):
    serializer_class = ProductCreateUpdateSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        user = self.request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            raise PermissionDenied("Доступ запрещён.")
        set_audit_user(user)  # для триггера аудита в БД
        serializer.save()
        # Аудит выполняется автоматически триггером trg_product_audit

    def create(self, request, *args, **kwargs):
        if request.user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        return super().create(request, *args, **kwargs)

class AdminProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'
    
    def get_serializer_class(self):
        if self.request.method == 'GET':
            return ProductWithRelationsSerializer
        return ProductCreateUpdateSerializer
    
    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Product.objects.all().select_related('categoryId', 'brandId').prefetch_related('productimage_set', 'productattribute_set')
        else:
            return Product.objects.none()
    
    def update(self, request, *args, **kwargs):
        user = request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        set_audit_user(user)  # для триггера аудита в БД
        response = super().update(request, *args, **kwargs)
        # Аудит выполняется автоматически триггером trg_product_audit
        return response

    def destroy(self, request, *args, **kwargs):
        user = request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        set_audit_user(user)  # для триггера аудита в БД
        response = super().destroy(request, *args, **kwargs)
        # Аудит выполняется автоматически триггером trg_product_audit
        return response


class ProductImageViewSet(viewsets.ModelViewSet):
    serializer_class = ProductImageSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            product_id = self.kwargs.get('product_pk')
            if product_id:
                return ProductImage.objects.filter(productId=product_id)
            return ProductImage.objects.all()
        return ProductImage.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            product_id = self.kwargs.get('product_pk')
            if product_id:
                product = Product.objects.get(productId=product_id)
                serializer.save(productId=product)
                log_audit(user, 'CREATE', 'productImage', get_pk(serializer.instance), old_values=None, new_values=model_to_log_dict(serializer.instance))
        else:
            raise PermissionDenied("Доступ запрещён.")

    def perform_update(self, serializer):
        user = self.request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            raise PermissionDenied("Доступ запрещён.")
        old_values = model_to_log_dict(serializer.instance)
        super().perform_update(serializer)
        log_audit(user, 'UPDATE', 'productImage', get_pk(serializer.instance), old_values=old_values, new_values=model_to_log_dict(serializer.instance))

    def perform_destroy(self, instance):
        user = self.request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            raise PermissionDenied("Доступ запрещён.")
        old_values = model_to_log_dict(instance)
        record_id = get_pk(instance)
        super().perform_destroy(instance)
        log_audit(user, 'DELETE', 'productImage', record_id, old_values=old_values, new_values=None)

class ProductAttributeViewSet(viewsets.ModelViewSet):
    serializer_class = ProductAttributeSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            product_id = self.kwargs.get('product_pk')
            if product_id:
                return ProductAttribute.objects.filter(productId=product_id)
            return ProductAttribute.objects.all()
        return ProductAttribute.objects.none()
    
    def perform_create(self, serializer):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            product_id = self.kwargs.get('product_pk')
            if product_id:
                product = Product.objects.get(productId=product_id)
                serializer.save(productId=product)
                log_audit(user, 'CREATE', 'productAttribute', get_pk(serializer.instance), old_values=None, new_values=model_to_log_dict(serializer.instance))
        else:
            raise PermissionDenied("Доступ запрещён.")

    def perform_update(self, serializer):
        user = self.request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            raise PermissionDenied("Доступ запрещён.")
        old_values = model_to_log_dict(serializer.instance)
        super().perform_update(serializer)
        log_audit(user, 'UPDATE', 'productAttribute', get_pk(serializer.instance), old_values=old_values, new_values=model_to_log_dict(serializer.instance))

    def perform_destroy(self, instance):
        user = self.request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            raise PermissionDenied("Доступ запрещён.")
        old_values = model_to_log_dict(instance)
        record_id = get_pk(instance)
        super().perform_destroy(instance)
        log_audit(user, 'DELETE', 'productAttribute', record_id, old_values=old_values, new_values=None)


class ProductImageUploadView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        user = request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Файл не предоставлен.'}, status=400)
        if not file.content_type.startswith('image/'):
            return Response({'detail': 'Файл должен быть изображением.'}, status=400)
        ext = os.path.splitext(file.name)[1] or '.jpg'
        filename = f"{uuid.uuid4().hex}{ext}"
        path = os.path.join('products', filename)
        saved_path = default_storage.save(path, file)
        url = f"{settings.MEDIA_URL}{saved_path}"
        return Response({'url': url}, status=201)


# Admin CRUD для категорий и брендов
class AdminCategoryListCreateView(generics.ListCreateAPIView):
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Category.objects.all()
        return Category.objects.none()

    def create(self, request, *args, **kwargs):
        if request.user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        response = super().create(request, *args, **kwargs)
        if response.status_code == 201 and hasattr(response, 'data') and response.data.get('categoryId'):
            instance = Category.objects.get(categoryId=response.data['categoryId'])
            log_audit(request.user, 'CREATE', 'category', get_pk(instance), old_values=None, new_values=model_to_log_dict(instance))
        return response


class AdminCategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]
    lookup_url_kwarg = 'pk'

    def get_queryset(self):
        if self.request.user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Category.objects.all()
        return Category.objects.none()

    def update(self, request, *args, **kwargs):
        if request.user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        instance = self.get_object()
        old_values = model_to_log_dict(instance)
        response = super().update(request, *args, **kwargs)
        if response.status_code == 200:
            instance.refresh_from_db()
            log_audit(request.user, 'UPDATE', 'category', get_pk(instance), old_values=old_values, new_values=model_to_log_dict(instance))
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        instance = self.get_object()
        old_values = model_to_log_dict(instance)
        record_id = get_pk(instance)
        response = super().destroy(request, *args, **kwargs)
        if response.status_code in (200, 204):
            log_audit(request.user, 'DELETE', 'category', record_id, old_values=old_values, new_values=None)
        return response


class AdminBrandListCreateView(generics.ListCreateAPIView):
    serializer_class = BrandSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Brand.objects.all()
        return Brand.objects.none()

    def create(self, request, *args, **kwargs):
        if request.user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        response = super().create(request, *args, **kwargs)
        if response.status_code == 201 and hasattr(response, 'data') and response.data.get('brandId'):
            instance = Brand.objects.get(brandId=response.data['brandId'])
            log_audit(request.user, 'CREATE', 'brand', get_pk(instance), old_values=None, new_values=model_to_log_dict(instance))
        return response


class AdminBrandDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = BrandSerializer
    permission_classes = [IsAuthenticated]
    lookup_url_kwarg = 'pk'

    def get_queryset(self):
        if self.request.user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Brand.objects.all()
        return Brand.objects.none()

    def update(self, request, *args, **kwargs):
        if request.user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        instance = self.get_object()
        old_values = model_to_log_dict(instance)
        response = super().update(request, *args, **kwargs)
        if response.status_code == 200:
            instance.refresh_from_db()
            log_audit(request.user, 'UPDATE', 'brand', get_pk(instance), old_values=old_values, new_values=model_to_log_dict(instance))
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        instance = self.get_object()
        old_values = model_to_log_dict(instance)
        record_id = get_pk(instance)
        response = super().destroy(request, *args, **kwargs)
        if response.status_code in (200, 204):
            log_audit(request.user, 'DELETE', 'brand', record_id, old_values=old_values, new_values=None)
        return response


class AdminUsersView(generics.ListAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName == 'Администратор':
            return User.objects.all().select_related('roleId')
        return User.objects.none()


class AdminUserCreateView(generics.CreateAPIView):
    serializer_class = UserCreateUpdateSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        if self.request.user.roleId.roleName != 'Администратор':
            raise PermissionDenied("Доступ запрещён.")
        serializer.save()
        log_audit(
            self.request.user, 'CREATE', 'user',
            get_pk(serializer.instance),
            old_values=None,
            new_values=model_to_log_dict(serializer.instance),
        )

    def create(self, request, *args, **kwargs):
        if request.user.roleId.roleName != 'Администратор':
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        return super().create(request, *args, **kwargs)


class RoleListView(generics.ListAPIView):
    serializer_class = RoleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName == 'Администратор':
            return Role.objects.all()
        return Role.objects.none()


class AdminUserDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = UserCreateUpdateSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName == 'Администратор':
            return User.objects.all().select_related('roleId')
        return User.objects.none()

    def update(self, request, *args, **kwargs):
        if request.user.roleId.roleName != 'Администратор':
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        instance = self.get_object()
        if instance.userId == request.user.userId and request.data.get('is_active') is False:
            return Response({'detail': 'Нельзя заблокировать свой аккаунт.'}, status=400)
        old_values = model_to_log_dict(instance)
        response = super().update(request, *args, **kwargs)
        if response.status_code == 200:
            instance.refresh_from_db()
            log_audit(request.user, 'UPDATE', 'user', get_pk(instance), old_values=old_values, new_values=model_to_log_dict(instance))
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.roleId.roleName != 'Администратор':
            return Response({'detail': 'Доступ запрещён.'}, status=403)
        instance = self.get_object()
        if instance.userId == request.user.userId:
            return Response({'detail': 'Нельзя удалить собственный аккаунт.'}, status=400)
        old_values = model_to_log_dict(instance)
        record_id = get_pk(instance)
        set_audit_user(request.user)  # для триггера аудита в БД
        response = super().destroy(request, *args, **kwargs)
        if response.status_code in (200, 204):
            log_audit(request.user, 'DELETE', 'user', record_id, old_values=old_values, new_values=None)
        return response

class AdminOrdersView(generics.ListAPIView):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Order.objects.all().select_related('userId', 'orderStatusId')
        return Order.objects.none()


class AdminOrderDetailView(generics.RetrieveUpdateAPIView):
    serializer_class = OrderDetailSerializer
    permission_classes = [IsAuthenticated]
    lookup_url_kwarg = 'pk'

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Order.objects.all().select_related('userId', 'orderStatusId', 'addressId').prefetch_related('orderitem_set__productId')
        return Order.objects.none()

    def put(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)

    def perform_update(self, serializer):
        set_audit_user(self.request.user)  # для триггера аудита в БД
        super().perform_update(serializer)
        # Аудит выполняется автоматически триггером trg_order_audit


class AdminOrderMarkPaidView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Order.objects.all().select_related('orderStatusId')
        return Order.objects.none()

    def post(self, request, pk, *args, **kwargs):
        try:
            order = self.get_queryset().get(orderId=pk)
            if order.paymentType not in [Order.PAYMENT_CARD, Order.PAYMENT_CASH]:
                return Response(
                    {'detail': 'Этот заказ не предназначен для оплаты при получении.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if order.paymentStatus == Order.PAYMENT_STATUS_PAID:
                return Response({'detail': 'Заказ уже отмечен как оплаченный.'}, status=status.HTTP_400_BAD_REQUEST)
            set_audit_user(request.user)  # для триггера аудита в БД
            order.paymentStatus = Order.PAYMENT_STATUS_PAID
            order.save(update_fields=['paymentStatus'])
            # Аудит выполняется автоматически триггером trg_order_audit
        except Order.DoesNotExist:
            return Response({'detail': 'Заказ не найден.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.exception("AdminOrderMarkPaid failed")
            return Response({'detail': 'Не удалось обновить статус оплаты.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        serializer = OrderDetailSerializer(order)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AdminOrderStatusListView(generics.ListAPIView):
    serializer_class = OrderStatusSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            return OrderStatus.objects.all()
        return OrderStatus.objects.none()


class AdminAuditLogListView(generics.ListAPIView):
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName == 'Администратор':
            return AuditLog.objects.all().select_related('userId').order_by('-createdAt')
        return AuditLog.objects.none()


class AdminReviewListView(generics.ListAPIView):
    serializer_class = AdminReviewSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Review.objects.all().select_related('userId', 'productId').order_by('-createdAt')
        return Review.objects.none()


class AdminReviewDetailView(generics.RetrieveDestroyAPIView):
    serializer_class = AdminReviewSerializer
    permission_classes = [IsAuthenticated]
    lookup_url_kwarg = 'pk'

    def get_queryset(self):
        user = self.request.user
        if user.roleId.roleName in ['Администратор', 'Менеджер']:
            return Review.objects.all().select_related('userId', 'productId')
        return Review.objects.none()

    def perform_destroy(self, instance):
        record_id = get_pk(instance)
        old_values = model_to_log_dict(instance)
        super().perform_destroy(instance)
        log_audit(
            self.request.user, 'DELETE', 'review',
            record_id, old_values=old_values, new_values=None,
        )

class AdminAnalyticsSalesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=status.HTTP_403_FORBIDDEN)

        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        group_by = request.query_params.get('group_by', 'day')  
        export_format = request.query_params.get('export') 
        
        orders_qs = Order.objects.filter(paymentStatus=Order.PAYMENT_STATUS_PAID)

        if date_from:
            try:
                date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d')
                orders_qs = orders_qs.filter(createdAt__date__gte=date_from_parsed)
            except ValueError:
                pass
        if date_to:
            try:
                date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d')
                orders_qs = orders_qs.filter(createdAt__date__lte=date_to_parsed)
            except ValueError:
                pass

        total_stats = orders_qs.aggregate(
            total_revenue=Sum('total'),
            total_orders=Count('orderId'),
            avg_order_value=Avg('total')
        )

        # Данные из SQL-представления v_sales_report (сгруппированы по месяцам)
        # Фильтрация по дате применяется к представлению через WHERE
        sales_report_sql = 'SELECT "month", "orderCount", "revenue", "avgOrderTotal" FROM v_sales_report'
        sales_params = []
        sales_conditions = []
        if date_from:
            sales_conditions.append('"month" >= %s')
            sales_params.append(date_from)
        if date_to:
            sales_conditions.append('"month" <= %s')
            sales_params.append(date_to)
        if sales_conditions:
            sales_report_sql += ' WHERE ' + ' AND '.join(sales_conditions)
        sales_report_sql += ' ORDER BY "month" DESC'

        with connection.cursor() as cursor:
            cursor.execute(sales_report_sql, sales_params)
            sales_report_rows = cursor.fetchall()

        # Дополнительная группировка по дням/неделям (при необходимости)
        if group_by == 'month':
            trunc_func = TruncMonth('createdAt')
        elif group_by == 'week':
            trunc_func = TruncWeek('createdAt')
        else:
            trunc_func = TruncDate('createdAt')

        sales_by_period = orders_qs.annotate(
            period=trunc_func
        ).values('period').annotate(
            revenue=Sum('total'),
            orders_count=Count('orderId'),
            avg_check=Avg('total')
        ).order_by('period')

        # Продажи по категориям
        sales_by_category = OrderItem.objects.filter(
            orderId__in=orders_qs
        ).values(
            category_name=F('productId__categoryId__categoryName')
        ).annotate(
            total_sold=Sum('quantity'),
            revenue=Sum(F('quantity') * F('unitPrice'))
        ).order_by('-revenue')[:10]

        # Продажи по брендам
        sales_by_brand = OrderItem.objects.filter(
            orderId__in=orders_qs
        ).values(
            brand_name=F('productId__brandId__brandName')
        ).annotate(
            total_sold=Sum('quantity'),
            revenue=Sum(F('quantity') * F('unitPrice'))
        ).order_by('-revenue')[:10]

        # Распределение по статусам заказов (все заказы, не только оплаченные)
        all_orders_qs = Order.objects.all()
        if date_from:
            try:
                date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d')
                all_orders_qs = all_orders_qs.filter(createdAt__date__gte=date_from_parsed)
            except ValueError:
                pass
        if date_to:
            try:
                date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d')
                all_orders_qs = all_orders_qs.filter(createdAt__date__lte=date_to_parsed)
            except ValueError:
                pass

        orders_by_status = all_orders_qs.values(
            status_name=F('orderStatusId__orderStatusName')
        ).annotate(
            count=Count('orderId')
        ).order_by('-count')

        # Распределение по способам оплаты
        orders_by_payment = all_orders_qs.values('paymentType').annotate(
            count=Count('orderId')
        ).order_by('-count')

        # Экспорт в файл если указан формат
        if export_format in ['csv', 'excel']:
            # Данные по дням для экспорта
            sales_data = orders_qs.annotate(
                date=TruncDate('createdAt')
            ).values('date').annotate(
                revenue=Sum('total'),
                orders_count=Count('orderId'),
                avg_check=Avg('total')
            ).order_by('date')

            rows = [
                {
                    'Дата': item['date'].strftime('%Y-%m-%d') if item['date'] else '',
                    'Выручка (₽)': float(item['revenue'] or 0),
                    'Количество заказов': item['orders_count'],
                    'Средний чек (₽)': round(float(item['avg_check'] or 0), 2),
                }
                for item in sales_data
            ]

            # Добавляем итоговую строку
            total = orders_qs.aggregate(
                total_revenue=Sum('total'),
                total_orders=Count('orderId'),
                avg_check=Avg('total')
            )
            rows.append({
                'Дата': 'ИТОГО',
                'Выручка (₽)': float(total['total_revenue'] or 0),
                'Количество заказов': total['total_orders'] or 0,
                'Средний чек (₽)': round(float(total['avg_check'] or 0), 2),
            })

            filename = f"sales_report_{date_from or 'all'}_{date_to or 'all'}"
            return self._create_export_response(rows, export_format, filename)

        return Response({
            'summary': {
                'total_revenue': float(total_stats['total_revenue'] or 0),
                'total_orders': total_stats['total_orders'] or 0,
                'avg_order_value': float(total_stats['avg_order_value'] or 0),
            },
            'sales_by_period': [
                {
                    'period': item['period'].strftime('%Y-%m-%d') if item['period'] else None,
                    'revenue': float(item['revenue'] or 0),
                    'orders_count': item['orders_count'],
                    'avg_check': float(item['avg_check'] or 0),
                }
                for item in sales_by_period
            ],
            'sales_by_category': [
                {
                    'category': item['category_name'] or 'Без категории',
                    'total_sold': item['total_sold'],
                    'revenue': float(item['revenue'] or 0),
                }
                for item in sales_by_category
            ],
            'sales_by_brand': [
                {
                    'brand': item['brand_name'] or 'Без бренда',
                    'total_sold': item['total_sold'],
                    'revenue': float(item['revenue'] or 0),
                }
                for item in sales_by_brand
            ],
            'orders_by_status': [
                {
                    'status': item['status_name'] or 'Неизвестно',
                    'count': item['count'],
                }
                for item in orders_by_status
            ],
            'orders_by_payment': [
                {
                    'payment_type': item['paymentType'],
                    'count': item['count'],
                }
                for item in orders_by_payment
            ],
            # Данные из SQL-представления v_sales_report (помесячная сводка)
            'monthly_report': [
                {
                    'month': str(row[0]) if row[0] else None,
                    'orderCount': row[1],
                    'revenue': float(row[2]) if row[2] else 0,
                    'avgOrderTotal': float(row[3]) if row[3] else 0,
                }
                for row in sales_report_rows
            ],
        })

    def _create_export_response(self, rows, export_format, filename):
        print(f">>> SALES EXPORT: format={export_format}, OPENPYXL={OPENPYXL_AVAILABLE} <<<")
        if not rows:
            rows = [{'Нет данных': ''}]

        if export_format == 'excel' and OPENPYXL_AVAILABLE:
            print(">>> CREATING EXCEL FILE <<<")
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчёт"

            headers = list(rows[0].keys())
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                    cell.border = thin_border

            for col_idx, header in enumerate(headers, 1):
                max_length = max(len(str(header)), max(len(str(row[header])) for row in rows))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response

        output = io.StringIO()
        if rows:
            fieldnames = list(rows[0].keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        response = HttpResponse(
            output.getvalue().encode('utf-8-sig'),
            content_type='text/csv; charset=utf-8-sig'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response


class AdminAnalyticsProductsView(APIView):
    """Аналитика популярности товаров."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=status.HTTP_403_FORBIDDEN)

        # Получаем параметры фильтрации
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        limit = int(request.query_params.get('limit', 20))
        export_format = request.query_params.get('export')  # csv, excel

        # Базовый queryset - только оплаченные заказы
        orders_qs = Order.objects.filter(paymentStatus=Order.PAYMENT_STATUS_PAID)

        # Применяем фильтры по дате
        if date_from:
            try:
                date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d')
                orders_qs = orders_qs.filter(createdAt__date__gte=date_from_parsed)
            except ValueError:
                pass
        if date_to:
            try:
                date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d')
                orders_qs = orders_qs.filter(createdAt__date__lte=date_to_parsed)
            except ValueError:
                pass

        # Топ продаваемых товаров — из SQL-представления v_popular_products
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "productId", "productName", "categoryName", "totalSold", "totalRevenue", "avgRating" '
                'FROM v_popular_products LIMIT %s',
                [limit]
            )
            popular_rows = cursor.fetchall()

        top_products = [
            {
                'product_id': row[0],
                'product_name': row[1],
                'category_name': row[2],
                'brand_name': '',
                'total_sold': row[3],
                'revenue': float(row[4]) if row[4] else 0,
                'orders_count': 0,
            }
            for row in popular_rows
        ]

        # Товары с наибольшей выручкой — из того же представления, отсортированы по выручке
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "productId", "productName", "categoryName", "totalSold", "totalRevenue" '
                'FROM v_popular_products ORDER BY "totalRevenue" DESC LIMIT %s',
                [limit]
            )
            revenue_rows = cursor.fetchall()

        top_revenue_products = [
            {
                'product_id': row[0],
                'product_name': row[1],
                'category_name': row[2],
                'total_sold': row[3],
                'revenue': float(row[4]) if row[4] else 0,
            }
            for row in revenue_rows
        ]

        # Товары с наивысшим рейтингом — из SQL-представления v_product_catalog
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "productId", "productName", "categoryName", "avgRating", "reviewCount" '
                'FROM v_product_catalog WHERE "reviewCount" > 0 '
                'ORDER BY "avgRating" DESC, "reviewCount" DESC LIMIT %s',
                [limit]
            )
            rated_rows = cursor.fetchall()

        top_rated_list = [
            {
                'product_id': row[0],
                'product_name': row[1],
                'category': row[2],
                'avg_rating': float(row[3]) if row[3] else 0,
                'reviews_count': row[4],
            }
            for row in rated_rows
        ]

        # Товары с низким остатком (менее 5 штук)
        low_stock_products = Product.objects.filter(
            quantity__lt=5
        ).order_by('quantity')[:20]

        low_stock_list = [
            {
                'product_id': p.productId,
                'product_name': p.productName,
                'category': p.categoryId.categoryName if p.categoryId else None,
                'quantity': p.quantity,
                'price': float(p.price),
            }
            for p in low_stock_products
        ]

        # Экспорт в файл если указан формат
        if export_format in ['csv', 'excel']:
            products_data = OrderItem.objects.filter(
                orderId__in=orders_qs
            ).values(
                product_name=F('productId__productName'),
                category_name=F('productId__categoryId__categoryName'),
                brand_name=F('productId__brandId__brandName'),
            ).annotate(
                total_sold=Sum('quantity'),
                revenue=Sum(F('quantity') * F('unitPrice')),
                orders_count=Count('orderId', distinct=True)
            ).order_by('-total_sold')

            rows = [
                {
                    'Товар': item['product_name'],
                    'Категория': item['category_name'] or 'Без категории',
                    'Бренд': item['brand_name'] or 'Без бренда',
                    'Продано (шт.)': item['total_sold'],
                    'Выручка (₽)': float(item['revenue'] or 0),
                    'Заказов': item['orders_count'],
                }
                for item in products_data
            ]

            filename = f"products_report_{date_from or 'all'}_{date_to or 'all'}"
            return self._create_export_response(rows, export_format, filename)

        return Response({
            'top_selling_products': [
                {
                    'product_id': item['product_id'],
                    'product_name': item['product_name'],
                    'category': item['category_name'] or 'Без категории',
                    'brand': item['brand_name'] or 'Без бренда',
                    'total_sold': item['total_sold'],
                    'revenue': float(item['revenue'] or 0),
                    'orders_count': item['orders_count'],
                }
                for item in top_products
            ],
            'top_revenue_products': [
                {
                    'product_id': item['product_id'],
                    'product_name': item['product_name'],
                    'category': item['category_name'] or 'Без категории',
                    'total_sold': item['total_sold'],
                    'revenue': float(item['revenue'] or 0),
                }
                for item in top_revenue_products
            ],
            'top_rated_products': top_rated_list,
            'low_stock_products': low_stock_list,
        })

    def _create_export_response(self, rows, export_format, filename):
        """Создаёт HTTP response с файлом для экспорта."""
        if not rows:
            rows = [{'Нет данных': ''}]

        if export_format == 'excel' and OPENPYXL_AVAILABLE:
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчёт"

            headers = list(rows[0].keys())
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                    cell.border = thin_border

            for col_idx, header in enumerate(headers, 1):
                max_length = max(len(str(header)), max(len(str(row[header])) for row in rows))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response

        output = io.StringIO()
        if rows:
            fieldnames = list(rows[0].keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        response = HttpResponse(
            output.getvalue().encode('utf-8-sig'),
            content_type='text/csv; charset=utf-8-sig'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response


class AdminPriceAdjustmentView(APIView):
    """Пакетное изменение цен по категории через хранимую процедуру sp_adjust_prices_by_category."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=status.HTTP_403_FORBIDDEN)

        category_id = request.data.get('categoryId')
        percent_change = request.data.get('percentChange')

        if category_id is None or percent_change is None:
            return Response(
                {'detail': 'Укажите categoryId и percentChange.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            category_id = int(category_id)
            percent_change = float(percent_change)
        except (ValueError, TypeError):
            return Response(
                {'detail': 'categoryId должен быть числом, percentChange — числом.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            set_audit_user(user)  # для триггеров аудита на product
            with connection.cursor() as cursor:
                cursor.execute(
                    'CALL sp_adjust_prices_by_category(%s, %s)',
                    [category_id, percent_change]
                )
            return Response({
                'message': f'Цены в категории обновлены на {percent_change:+.2f}%.'
            })
        except Exception as e:
            logger.exception("Ошибка изменения цен")
            error_msg = str(e)
            if 'не найдена' in error_msg or 'does not exist' in error_msg.lower():
                detail = 'Указанная категория не найдена.'
            elif 'Нет товаров' in error_msg:
                detail = 'В указанной категории нет товаров.'
            else:
                detail = 'Ошибка при изменении цен. Попробуйте позже.'
            return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)


class AdminUserActivityView(APIView):
    """Активность пользователей из SQL-представления v_user_activity."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=status.HTTP_403_FORBIDDEN)

        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "userId", "fullName", "email", "roleName", "orderCount", '
                '"totalSpent", "reviewCount", "registeredAt" '
                'FROM v_user_activity ORDER BY "totalSpent" DESC LIMIT 50'
            )
            rows = cursor.fetchall()

        data = [
            {
                'userId': row[0],
                'fullName': row[1],
                'email': row[2],
                'role': row[3],
                'orderCount': row[4],
                'totalSpent': float(row[5]) if row[5] else 0,
                'reviewCount': row[6],
                'registeredAt': row[7].isoformat() if row[7] else None,
            }
            for row in rows
        ]
        return Response(data)


class AdminAnalyticsExportView(APIView):
    """Экспорт отчётов в CSV и Excel."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        logger.info("AdminAnalyticsExportView.get() called")
        user = request.user
        if user.roleId.roleName not in ['Администратор', 'Менеджер']:
            return Response({'detail': 'Доступ запрещён.'}, status=status.HTTP_403_FORBIDDEN)

        report_type = request.query_params.get('report', 'sales')  
        export_format = request.query_params.get('format', 'csv')  
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        # Базовый queryset
        orders_qs = Order.objects.filter(paymentStatus=Order.PAYMENT_STATUS_PAID)

        if date_from:
            try:
                date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d')
                orders_qs = orders_qs.filter(createdAt__date__gte=date_from_parsed)
            except ValueError:
                pass
        if date_to:
            try:
                date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d')
                orders_qs = orders_qs.filter(createdAt__date__lte=date_to_parsed)
            except ValueError:
                pass

        if report_type == 'products':
            return self._export_products_report(orders_qs, export_format, date_from, date_to)
        else:
            return self._export_sales_report(orders_qs, export_format, date_from, date_to)

    def _export_sales_report(self, orders_qs, export_format, date_from, date_to):
        """Экспорт отчёта о продажах."""
        # Данные по дням
        sales_data = orders_qs.annotate(
            date=TruncDate('createdAt')
        ).values('date').annotate(
            revenue=Sum('total'),
            orders_count=Count('orderId'),
            avg_check=Avg('total')
        ).order_by('date')

        rows = [
            {
                'Дата': item['date'].strftime('%Y-%m-%d') if item['date'] else '',
                'Выручка (₽)': float(item['revenue'] or 0),
                'Количество заказов': item['orders_count'],
                'Средний чек (₽)': round(float(item['avg_check'] or 0), 2),
            }
            for item in sales_data
        ]

        # Добавляем итоговую строку
        total = orders_qs.aggregate(
            total_revenue=Sum('total'),
            total_orders=Count('orderId'),
            avg_check=Avg('total')
        )
        rows.append({
            'Дата': 'ИТОГО',
            'Выручка (₽)': float(total['total_revenue'] or 0),
            'Количество заказов': total['total_orders'] or 0,
            'Средний чек (₽)': round(float(total['avg_check'] or 0), 2),
        })

        filename = f"sales_report_{date_from or 'all'}_{date_to or 'all'}"
        return self._create_response(rows, export_format, filename)

    def _export_products_report(self, orders_qs, export_format, date_from, date_to):
        """Экспорт отчёта о популярности товаров."""
        products_data = OrderItem.objects.filter(
            orderId__in=orders_qs
        ).values(
            product_name=F('productId__productName'),
            category_name=F('productId__categoryId__categoryName'),
            brand_name=F('productId__brandId__brandName'),
        ).annotate(
            total_sold=Sum('quantity'),
            revenue=Sum(F('quantity') * F('unitPrice')),
            orders_count=Count('orderId', distinct=True)
        ).order_by('-total_sold')

        rows = [
            {
                'Товар': item['product_name'],
                'Категория': item['category_name'] or 'Без категории',
                'Бренд': item['brand_name'] or 'Без бренда',
                'Продано (шт.)': item['total_sold'],
                'Выручка (₽)': float(item['revenue'] or 0),
                'Заказов': item['orders_count'],
            }
            for item in products_data
        ]

        filename = f"products_report_{date_from or 'all'}_{date_to or 'all'}"
        return self._create_response(rows, export_format, filename)

    def _create_response(self, rows, export_format, filename):
        """Создаёт HTTP response с файлом."""
        if not rows:
            rows = [{'Нет данных': ''}]

        if export_format == 'excel':
            return self._create_excel_response(rows, filename)
        else:
            return self._create_csv_response(rows, filename)

    def _create_csv_response(self, rows, filename):
        """Создаёт CSV файл."""
        output = io.StringIO()
        if rows:
            fieldnames = list(rows[0].keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        response = HttpResponse(
            output.getvalue().encode('utf-8-sig'),  # BOM для корректного отображения в Excel
            content_type='text/csv; charset=utf-8-sig'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response

    def _create_excel_response(self, rows, filename):
        """Создаёт Excel файл (XLSX) используя openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Отчёт"

            if rows:
                # Заголовки
                headers = list(rows[0].keys())
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                # Данные
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=row_data[header])
                        cell.border = thin_border
                        if isinstance(row_data[header], (int, float)):
                            cell.alignment = Alignment(horizontal='right')

                # Автоширина колонок
                for col_idx, header in enumerate(headers, 1):
                    max_length = len(str(header))
                    for row_data in rows:
                        cell_length = len(str(row_data[header]))
                        if cell_length > max_length:
                            max_length = cell_length
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response

        except ImportError:
            # Если openpyxl не установлен, возвращаем CSV
            return self._create_csv_response(rows, filename)


EXPORT_TABLE_CONFIG = {
    'product': {
        'model': Product,
        'fields': ['productId', 'productName', 'productDescription', 'categoryId_id', 'brandId_id', 'price', 'ageRating', 'quantity', 'weightKg', 'dimensions'],
        'headers': ['productId', 'productName', 'productDescription', 'categoryId', 'brandId', 'price', 'ageRating', 'quantity', 'weightKg', 'dimensions'],
        'db_table': 'product',
    },
    'category': {
        'model': Category,
        'fields': ['categoryId', 'categoryName', 'categoryDescription'],
        'headers': ['categoryId', 'categoryName', 'categoryDescription'],
        'db_table': 'category',
    },
    'brand': {
        'model': Brand,
        'fields': ['brandId', 'brandName', 'brandDescription', 'brandCountry'],
        'headers': ['brandId', 'brandName', 'brandDescription', 'brandCountry'],
        'db_table': 'brand',
    },
    'order': {
        'model': Order,
        'fields': ['orderId', 'userId_id', 'orderStatusId_id', 'total', 'addressId_id', 'deliveryType', 'paymentType', 'paymentStatus', 'note', 'createdAt'],
        'headers': ['orderId', 'userId', 'orderStatusId', 'total', 'addressId', 'deliveryType', 'paymentType', 'paymentStatus', 'note', 'createdAt'],
        'db_table': '"order"',
    },
    'user': {
        'model': User,
        'fields': ['userId', 'lastName', 'firstName', 'middleName', 'email', 'roleId_id', 'phone', 'birthDate', 'createdAt'],
        'headers': ['userId', 'lastName', 'firstName', 'middleName', 'email', 'roleId', 'phone', 'birthDate', 'createdAt'],
        'db_table': '"user"',
    },
    'review': {
        'model': Review,
        'fields': ['reviewId', 'productId_id', 'userId_id', 'rating', 'reviewText', 'createdAt', 'updatedAt'],
        'headers': ['reviewId', 'productId', 'userId', 'rating', 'reviewText', 'createdAt', 'updatedAt'],
        'db_table': 'review',
    },
    'auditlog': {
        'model': AuditLog,
        'fields': ['auditLogId', 'userId_id', 'action', 'tableName', 'recordId', 'oldValues', 'newValues', 'createdAt'],
        'headers': ['auditLogId', 'userId', 'action', 'tableName', 'recordId', 'oldValues', 'newValues', 'createdAt'],
        'db_table': '"auditLog"',
    },
}

IMPORT_TABLE_CONFIG = {
    'product': {
        'model': Product,
        'fields_map': {
            'productName': 'productName',
            'productDescription': 'productDescription',
            'categoryId': 'categoryId_id',
            'brandId': 'brandId_id',
            'price': 'price',
            'ageRating': 'ageRating',
            'quantity': 'quantity',
            'weightKg': 'weightKg',
            'dimensions': 'dimensions',
        },
        'required': ['productName', 'categoryId', 'brandId', 'price'],
    },
    'category': {
        'model': Category,
        'fields_map': {
            'categoryName': 'categoryName',
            'categoryDescription': 'categoryDescription',
        },
        'required': ['categoryName'],
    },
    'brand': {
        'model': Brand,
        'fields_map': {
            'brandName': 'brandName',
            'brandDescription': 'brandDescription',
            'brandCountry': 'brandCountry',
        },
        'required': ['brandName'],
    },
}


class AdminDataExportView(APIView):
    """Экспорт данных таблиц в CSV или SQL формат."""
    permission_classes = [IsAuthenticated]

    def perform_content_negotiation(self, request, force=False):
        """Возвращает файл (HttpResponse), а не DRF Response — пропускаем строгую проверку."""
        from rest_framework.renderers import JSONRenderer
        return (JSONRenderer(), JSONRenderer.media_type)

    def get(self, request):
        user = request.user
        if not hasattr(user, 'roleId') or user.roleId.roleName != 'Администратор':
            return Response({'detail': 'Доступ запрещён.'}, status=403)

        table = request.query_params.get('table', '')
        fmt = request.query_params.get('file_format', 'csv')

        if table not in EXPORT_TABLE_CONFIG:
            return Response({'detail': f'Неизвестная таблица: {table}'}, status=400)

        config = EXPORT_TABLE_CONFIG[table]
        model = config['model']
        fields = config['fields']
        headers = config['headers']
        db_table = config['db_table']

        queryset = model.objects.all().order_by(model._meta.pk.name)
        rows = queryset.values_list(*fields)

        if fmt == 'sql':
            output = io.StringIO()
            output.write(f'-- Экспорт таблицы {db_table}\n')
            output.write(f'-- Дата: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
            output.write(f'-- Записей: {len(rows)}\n\n')

            for row in rows:
                values = []
                for val in row:
                    if val is None:
                        values.append('NULL')
                    elif isinstance(val, (int, float, Decimal)):
                        values.append(str(val))
                    elif isinstance(val, (dict, list)):
                        s = _json.dumps(val, ensure_ascii=False).replace("'", "''")
                        values.append(f"'{s}'")
                    elif isinstance(val, datetime):
                        values.append(f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'")
                    elif hasattr(val, 'isoformat'):
                        values.append(f"'{val.isoformat()}'")
                    else:
                        s = str(val).replace("'", "''")
                        values.append(f"'{s}'")

                cols = ', '.join(f'"{h}"' for h in headers)
                vals = ', '.join(values)
                output.write(f'INSERT INTO {db_table} ({cols}) VALUES ({vals});\n')

            response = HttpResponse(output.getvalue(), content_type='text/sql; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{table}_export.sql"'
            return response

        else:
            output = io.StringIO()
            writer = csv.writer(output, quoting=csv.QUOTE_ALL)
            writer.writerow(headers)
            for row in rows:
                processed = []
                for val in row:
                    if val is None:
                        processed.append('')
                    elif isinstance(val, (dict, list)):
                        processed.append(_json.dumps(val, ensure_ascii=False))
                    else:
                        processed.append(str(val))
                writer.writerow(processed)

            response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="{table}_export.csv"'
            return response


class AdminDataImportView(APIView):
    """Импорт данных из CSV."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        user = request.user
        if not hasattr(user, 'roleId') or user.roleId.roleName != 'Администратор':
            return Response({'detail': 'Доступ запрещён.'}, status=403)

        table = request.data.get('table', '')
        csv_file = request.FILES.get('file')

        if not csv_file:
            return Response({'detail': 'CSV-файл не предоставлен.'}, status=400)

        if table not in IMPORT_TABLE_CONFIG:
            return Response({'detail': f'Импорт в таблицу «{table}» не поддерживается.'}, status=400)

        config = IMPORT_TABLE_CONFIG[table]
        model = config['model']
        fields_map = config['fields_map']
        required = config['required']

        try:
            content = csv_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            csv_headers = reader.fieldnames or []

            missing_required = [f for f in required if f not in csv_headers]
            if missing_required:
                return Response({
                    'detail': f'В CSV отсутствуют обязательные столбцы: {", ".join(missing_required)}'
                }, status=400)

            created_count = 0
            errors = []

            for i, row in enumerate(reader, start=2):
                try:
                    obj_data = {}
                    for csv_col, model_field in fields_map.items():
                        if csv_col in row and row[csv_col].strip() != '':
                            obj_data[model_field] = row[csv_col].strip()

                    missing_in_row = [f for f in required if f not in row or row[f].strip() == '']
                    if missing_in_row:
                        errors.append(f'Строка {i}: пустые обязательные поля: {", ".join(missing_in_row)}')
                        continue

                    model.objects.create(**obj_data)
                    created_count += 1
                except Exception as e:
                    errors.append(f'Строка {i}: {str(e)}')

            detail = f'Создано записей: {created_count}.'
            if errors:
                detail += f' Ошибок: {len(errors)}. Первые ошибки: ' + '; '.join(errors[:5])

            return Response({'detail': detail}, status=200 if created_count > 0 else 400)

        except UnicodeDecodeError:
            return Response({'detail': 'Невозможно прочитать файл. Убедитесь, что он в формате CSV (UTF-8).'}, status=400)
        except Exception as e:
            return Response({'detail': f'Ошибка импорта: {str(e)}'}, status=500)


class AdminBackupListView(APIView):
    """Список резервных копий и создание новой."""
    permission_classes = [IsAuthenticated]

    def _check_admin(self, user):
        if user.roleId.roleName != 'Администратор':
            return Response({'detail': 'Доступно только администратору.'}, status=status.HTTP_403_FORBIDDEN)
        return None

    def get(self, request):
        """Список существующих бэкапов."""
        denied = self._check_admin(request.user)
        if denied:
            return denied

        backup_dir = Path(settings.BACKUP_DIR)
        if not backup_dir.exists():
            return Response({'backups': [], 'backup_dir': str(backup_dir)})

        backups = []
        for f in sorted(backup_dir.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
            if f.is_file() and f.suffix in ('.backup', '.sql'):
                stat = f.stat()
                backups.append({
                    'filename': f.name,
                    'size': stat.st_size,
                    'sizeHuman': self._human_size(stat.st_size),
                    'createdAt': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    'format': 'custom' if f.suffix == '.backup' else 'sql',
                })

        return Response({
            'backups': backups,
            'backup_dir': str(backup_dir),
            'max_count': getattr(settings, 'BACKUP_MAX_COUNT', 10),
        })

    def post(self, request):
        """Создание новой резервной копии."""
        denied = self._check_admin(request.user)
        if denied:
            return denied

        fmt = request.data.get('format', 'custom')
        data_only = request.data.get('dataOnly', False)

        if fmt not in ('custom', 'sql'):
            return Response({'detail': 'Формат должен быть custom или sql.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from django.core.management import call_command
            from io import StringIO
            out = StringIO()
            call_command('backup_db', format=fmt, data_only=data_only, stdout=out)
            output = out.getvalue()

            # Парсим путь к файлу из вывода команды
            log_audit(request.user, 'CREATE', 'backup', 0,
                      old_values=None,
                      new_values={'format': fmt, 'data_only': data_only, 'output': output.strip()})

            return Response({
                'detail': 'Резервная копия успешно создана.',
                'output': output.strip(),
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.exception("Ошибка создания бэкапа")
            return Response({
                'detail': f'Ошибка создания резервной копии: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @staticmethod
    def _human_size(nbytes):
        for unit in ('Б', 'КБ', 'МБ', 'ГБ'):
            if abs(nbytes) < 1024:
                return f'{nbytes:.1f} {unit}'
            nbytes /= 1024
        return f'{nbytes:.1f} ТБ'


class AdminBackupDownloadView(APIView):
    """Скачивание файла бэкапа (поддерживает ?token= для прямой ссылки)."""
    permission_classes = [AllowAny]  # Проверяем токен вручную

    def get(self, request, filename):
        # Поддержка токена через query-параметр (для скачивания в новой вкладке)
        user = request.user
        if not user or not user.is_authenticated:
            token_key = request.query_params.get('token')
            if token_key:
                try:
                    token = Token.objects.get(key=token_key)
                    user = token.user
                except Token.DoesNotExist:
                    return Response({'detail': 'Недействительный токен.'}, status=status.HTTP_401_UNAUTHORIZED)
            else:
                return Response({'detail': 'Необходима авторизация.'}, status=status.HTTP_401_UNAUTHORIZED)

        if user.roleId.roleName != 'Администратор':
            return Response({'detail': 'Доступно только администратору.'}, status=status.HTTP_403_FORBIDDEN)

        backup_dir = Path(settings.BACKUP_DIR)
        filepath = backup_dir / filename

        # Защита от path traversal
        if not filepath.resolve().parent == backup_dir.resolve():
            return Response({'detail': 'Недопустимое имя файла.'}, status=status.HTTP_400_BAD_REQUEST)

        if not filepath.exists():
            return Response({'detail': 'Файл не найден.'}, status=status.HTTP_404_NOT_FOUND)

        return FileResponse(
            open(filepath, 'rb'),
            as_attachment=True,
            filename=filename,
        )


class AdminBackupDeleteView(APIView):
    """Удаление файла бэкапа."""
    permission_classes = [IsAuthenticated]

    def delete(self, request, filename):
        if request.user.roleId.roleName != 'Администратор':
            return Response({'detail': 'Доступно только администратору.'}, status=status.HTTP_403_FORBIDDEN)

        backup_dir = Path(settings.BACKUP_DIR)
        filepath = backup_dir / filename

        if not filepath.resolve().parent == backup_dir.resolve():
            return Response({'detail': 'Недопустимое имя файла.'}, status=status.HTTP_400_BAD_REQUEST)

        if not filepath.exists():
            return Response({'detail': 'Файл не найден.'}, status=status.HTTP_404_NOT_FOUND)

        filepath.unlink()
        log_audit(request.user, 'DELETE', 'backup', 0,
                  old_values={'filename': filename}, new_values=None)
        return Response({'detail': f'Бэкап «{filename}» удалён.'}, status=status.HTTP_200_OK)


class AdminBackupRestoreView(APIView):
    """Восстановление БД из бэкапа."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.user.roleId.roleName != 'Администратор':
            return Response({'detail': 'Доступно только администратору.'}, status=status.HTTP_403_FORBIDDEN)

        filename = request.data.get('filename')
        if not filename:
            return Response({'detail': 'Укажите имя файла (filename).'}, status=status.HTTP_400_BAD_REQUEST)

        backup_dir = Path(settings.BACKUP_DIR)
        filepath = backup_dir / filename

        if not filepath.resolve().parent == backup_dir.resolve():
            return Response({'detail': 'Недопустимое имя файла.'}, status=status.HTTP_400_BAD_REQUEST)

        if not filepath.exists():
            return Response({'detail': 'Файл не найден.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            from django.core.management import call_command
            from io import StringIO
            out = StringIO()
            err = StringIO()
            call_command('restore_db', str(filepath), no_confirm=True, stdout=out, stderr=err)

            log_audit(request.user, 'UPDATE', 'backup_restore', 0,
                      old_values=None,
                      new_values={'filename': filename})

            return Response({
                'detail': f'База данных восстановлена из «{filename}».',
            })
        except Exception as e:
            logger.exception("Ошибка восстановления БД")
            return Response({
                'detail': f'Ошибка восстановления: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
